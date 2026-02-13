"""DFS XLSX ingestion utilities for retrieval index construction."""

from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any

from .chroma_retriever import RetrievalDocument

HEADER_SANITIZER = re.compile(r"[^a-z0-9]+")
PREFERRED_TEXT_FIELDS = [
    "case_number",
    "treatment",
    "diagnosis",
    "health_plan",
    "coverage_type",
    "decision",
    "determination",
    "rationale",
    "description",
    "clinical_background",
    "reviewer_rationale",
]
PREFERRED_METADATA_FIELDS = [
    "case_number",
    "decision_year",
    "health_plan",
    "coverage_type",
    "treatment",
    "diagnosis",
]


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = HEADER_SANITIZER.sub("_", text).strip("_")
    return text


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def _pick_case_id(row_map: dict[str, str], row_index: int) -> str:
    for key in ("case_number", "case_no", "case"):
        value = row_map.get(key, "")
        if value:
            return value
    return f"dfs_row_{row_index}"


def _build_text(row_map: dict[str, str]) -> str:
    parts: list[str] = []
    used = set()

    for field in PREFERRED_TEXT_FIELDS:
        value = row_map.get(field, "")
        if value:
            parts.append(f"{field.replace('_', ' ')}: {value}")
            used.add(field)

    for field, value in row_map.items():
        if field in used or not value:
            continue
        parts.append(f"{field.replace('_', ' ')}: {value}")

    return "\n".join(parts).strip()


def _build_metadata(row_map: dict[str, str]) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    for field in PREFERRED_METADATA_FIELDS:
        value = row_map.get(field, "")
        if value:
            metadata[field] = value
    return metadata


def load_dfs_documents(xlsx_path: Path, limit: int | None = None) -> list[RetrievalDocument]:
    """Load DFS records from XLSX and convert to retrieval documents."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse DFS XLSX files. Install with `pip install openpyxl`."
        ) from exc

    if not xlsx_path.exists():
        raise FileNotFoundError(f"DFS XLSX not found: {xlsx_path}")

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    header_row: tuple[Any, ...] | None = None
    for candidate in rows:
        if not candidate:
            continue
        if any(cell is not None and str(cell).strip() for cell in candidate):
            header_row = candidate
            break

    if header_row is None:
        return []

    headers = [_normalize_header(cell) for cell in header_row]
    documents: list[RetrievalDocument] = []

    for row_index, row in enumerate(rows, start=2):
        if limit is not None and len(documents) >= limit:
            break

        row_map: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = _clean_cell(row[idx] if idx < len(row) else "")
            if value:
                row_map[header] = value

        if not row_map:
            continue

        case_id = _pick_case_id(row_map, row_index)
        text = _build_text(row_map)
        if not text:
            continue

        metadata = _build_metadata(row_map)
        metadata["source"] = "ny_dfs_external_appeals"
        metadata["row_index"] = row_index

        documents.append(
            RetrievalDocument(
                doc_id=case_id,
                text=text,
                metadata=metadata,
            )
        )

    workbook.close()
    return documents
